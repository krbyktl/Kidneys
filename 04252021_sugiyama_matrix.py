# -*- coding: utf-8 -*-
"""
Created on Sun Apr 25 19:40:56 2021

@author: leokt
"""
import pandas as pd
import numpy as np

#import data and clean up
fileloc = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\clean sugiyama kinase substrates.xlsx"
ks_df = pd.read_excel(fileloc)

ks_df.rename(columns = {'(-6)':'negsix',
                        '(-5)':'negfive',
                        '(-4)':'negfour',
                        '(-3)':'negthree',
                        '(-2)':'negtwo',
                        '(-1)':'negone',
                        '(+1)':'posone',
                        '(+2)':'postwo',
                        '(+3)':'posthree',
                        '(+4)':'posfour',
                        '(+5)':'posfive',
                        '(+6)':'possix',}, inplace = True)

#divide dataframe per kinase
kinases = ks_df['Kinase'].unique()

AA = ['A','C','D','E','F','G','H','I','J','K','L','M','N','P','Q','R','S','T','V','W','Y']
#pos = {-6, -5, -4, -3, -2, -1, 0, 1, 2, 3, 4, 5, 6}

kin_ind = list(range(len(kinases)))
freq_ind = list(range(len(kinases)))

for i in kin_ind:
    kin_ind[i] = (ks_df[ks_df['Kinase']==kinases[i]])

    #calculate frequencies in each position per kinase
    #initialize an empty matrix
    #rows
    M = 21
    #columns
    N= 13
    init = [ [ 0 for i in range(M) ] for j in range(N) ]
    
    ns = (kin_ind[i].groupby(['negsix']).size()/kin_ind[i].groupby(['negsix']).size().sum()).to_dict()
    nf = (kin_ind[i].groupby(['negfive']).size()/kin_ind[i].groupby(['negfive']).size().sum()).to_dict()
    nfr = (kin_ind[i].groupby(['negfour']).size()/kin_ind[i].groupby(['negfour']).size().sum()).to_dict()
    nt = (kin_ind[i].groupby(['negthree']).size()/kin_ind[i].groupby(['negthree']).size().sum()).to_dict()
    ntw = (kin_ind[i].groupby(['negtwo']).size()/kin_ind[i].groupby(['negtwo']).size().sum()).to_dict()
    no = (kin_ind[i].groupby(['negone']).size()/kin_ind[i].groupby(['negone']).size().sum()).to_dict()
    zero = (kin_ind[i].groupby([0]).size()/kin_ind[i].groupby([0]).size().sum()).to_dict()
    o = (kin_ind[i].groupby(['posone']).size()/kin_ind[i].groupby(['posone']).size().sum()).to_dict()
    tw = (kin_ind[i].groupby(['postwo']).size()/kin_ind[i].groupby(['postwo']).size().sum()).to_dict()
    t = (kin_ind[i].groupby(['posthree']).size()/kin_ind[i].groupby(['posthree']).size().sum()).to_dict()
    fr = (kin_ind[i].groupby(['posfour']).size()/kin_ind[i].groupby(['posfour']).size().sum()).to_dict()
    f = (kin_ind[i].groupby(['posfive']).size()/kin_ind[i].groupby(['posfive']).size().sum()).to_dict()
    s = (kin_ind[i].groupby(['possix']).size()/kin_ind[i].groupby(['possix']).size().sum()).to_dict()
    freq = [ns,nf,nfr,nt,ntw,no,zero,o,tw,t,fr,f,s]
    
    for j in freq: 
        for k in range(M):
            if AA[k] not in j:
                j[AA[k]] = 0

#map frequencies into a 21x13 matrix 
    for l in range(21): 
        for m in range(13):
            init[m][l] = freq[m][AA[l]]      
    print(init)
    freq_ind[i] =  np.transpose(np.asarray(init))
    
#exclude "J" and 0 position entry to convert to 20x12 matrix
freq_array_mod = list(range(len(kinases)))
for n in range(len(freq_ind)):
    mod_array = np.delete(freq_ind[n],8,0)
    mod_array = np.delete(mod_array,6,1)
    freq_array_mod[n] = mod_array

#placeholder dictionary
kinase_freq_dict = {k:v for k,v in zip(kinases,freq_array_mod)}

#introduce reference probabilities
fileloc1 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\human PPSP background.xlsx"
Sbkgd_df = pd.read_excel(fileloc1, 'S-center')
Tbkgd_df = pd.read_excel(fileloc1, 'T-center')
#Ybkgd_df = pd.read_excel(fileloc1, 'Y-center')
Sbkgd_df = Sbkgd_df.drop([20,21])
Sbkgd_df = Sbkgd_df.drop(['Position:',0,-7,7], axis=1)
Tbkgd_df = Tbkgd_df.drop([20,21])
Tbkgd_df = Tbkgd_df.drop(['Position:',0,-7,7], axis=1)
Sbkgd_array = Sbkgd_df.to_numpy()
Tbkgd_array = Tbkgd_df.to_numpy()
#average S and T backgrounds
STbkgd_array = np.mean([Sbkgd_array,Tbkgd_array],axis=0)/100

#convert observed frequencies to information content scores [matrix] per kinase
#IC(a,i)=P(a,i)*log((P(a,i)/Pref(a)),2)
IC_array = list(range(len(kinases)))
for b in IC_array:
    IC_array[b] = freq_array_mod[b]*np.log2(freq_array_mod[b]/STbkgd_array)


#extract clusters
fileloc2 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\TC rat IMCD clusters.xlsx"
IB_df = pd.read_excel(fileloc2,'I.B')
IA1_df = pd.read_excel(fileloc2,'I.A.1')
IA2a_df = pd.read_excel(fileloc2,'I.A.2.a')
IA2b_df = pd.read_excel(fileloc2,'I.A.2.b')
IA2c_df = pd.read_excel(fileloc2,'I.A.2.c')
IIA1a_df = pd.read_excel(fileloc2,'II.A.1.a')
IIA1b_df = pd.read_excel(fileloc2,'II.A.1.b')
IIB1a_df = pd.read_excel(fileloc2,'II.B.1.a')
IIB1b_df = pd.read_excel(fileloc2,'II.B.1.b')
IIIA_df = pd.read_excel(fileloc2,'III.A')
IIIB1_df = pd.read_excel(fileloc2,'III.B.1')
IIIB2_df = pd.read_excel(fileloc2,'III.B.2')
IVA_df = pd.read_excel(fileloc2,'IV.A')
IVB_df = pd.read_excel(fileloc2,'IV.B')

clusters_name = ['IB','IA1','IA2a','IA2b','IA2c','IIA1a','IIA1b',
            'IIB1a','IIB1b','IIIA','IIIB1','IIIB2','IVA','IVB']

clus_ind = [IB_df,IA1_df,IA2a_df,IA2b_df,IA2c_df,IIA1a_df,IIA1b_df,
            IIB1a_df,IIB1b_df,IIIA_df,IIIB1_df,IIIB2_df,IVA_df,IVB_df]

#create frequency matrices
clusters = list(range(len(clusters)))
for h in clusters:
    #calculate frequencies in each position per kinase
    #initialize an empty matrix
    #rows
    M = 21
    #columns
    N = 13
    init = [ [ 0 for i in range(M) ] for j in range(N) ]
    
    ns1 = (clus_ind[h].groupby(['minus 6']).size()/clus_ind[h].groupby(['minus 6']).size().sum()).to_dict()
    nf1 = (clus_ind[h].groupby([' minus 5']).size()/clus_ind[h].groupby([' minus 5']).size().sum()).to_dict()
    nfr1 = (clus_ind[h].groupby(['minus 4']).size()/clus_ind[h].groupby(['minus 4']).size().sum()).to_dict()
    nt1 = (clus_ind[h].groupby(['minus 3']).size()/clus_ind[h].groupby(['minus 3']).size().sum()).to_dict()
    ntw1 = (clus_ind[h].groupby(['minus 2']).size()/clus_ind[h].groupby(['minus 2']).size().sum()).to_dict()
    no1 = (clus_ind[h].groupby(['minus 1']).size()/clus_ind[h].groupby(['minus 1']).size().sum()).to_dict()
    zero1 = (clus_ind[h].groupby(['zero']).size()/clus_ind[h].groupby(['zero']).size().sum()).to_dict()
    o1 = (clus_ind[h].groupby(['plus 1']).size()/clus_ind[h].groupby(['plus 1']).size().sum()).to_dict()
    tw1 = (clus_ind[h].groupby(['plus 2']).size()/clus_ind[h].groupby(['plus 2']).size().sum()).to_dict()
    t1 = (clus_ind[h].groupby(['plus 3']).size()/clus_ind[h].groupby(['plus 3']).size().sum()).to_dict()
    fr1 = (clus_ind[h].groupby(['plus 4']).size()/clus_ind[h].groupby(['plus 4']).size().sum()).to_dict()
    f1 = (clus_ind[h].groupby(['plus 5']).size()/clus_ind[h].groupby(['plus 5']).size().sum()).to_dict()
    s1 = (clus_ind[h].groupby(['plus 6']).size()/clus_ind[h].groupby(['plus 6']).size().sum()).to_dict()
    freq1 = [ns1,nf1,nfr1,nt1,ntw1,no1,zero1,o1,tw1,t1,fr1,f1,s1]
    
    for z in freq1: 
        for x in range(M):
            if AA[x] not in z:
                z[AA[x]] = 0       
#map frequencies into a 21x13 matrix 
    for l in range(21): 
        for m in range(13):
            init[m][l] = freq1[m][AA[l]]      
    print(init)
    clusters[h] =  np.transpose(np.asarray(init))

#exclude "J" and 0 position entry to convert to 20x12 matrix
cluster_array_mod = list(range(len(clusters)))
for n in range(len(clusters)):
    mod_array = np.delete(clusters[n],8,0)
    mod_array = np.delete(mod_array,6,1)
    cluster_array_mod[n] = mod_array


#import IMCD rat background
fileloc3 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\TC rat IMCD background.xlsx"
IMCDbkgd_df = pd.read_excel(fileloc3)
IMCDbkgd_array = (IMCDbkgd_df.to_numpy())/100

#create cluster IC matrices
#IC(a,i)=P(a,i)*log((P(a,i)/Pref(a)),2)
IC_cluster_array = list(range(len(clusters)))
for b in IC_cluster_array:
    IC_cluster_array[b] = cluster_array_mod[b]*np.log2(cluster_array_mod[b]/IMCDbkgd_array)

cluster_IC_dict = {k:v for k,v in zip(clusters_name,IC_cluster_array)}

#flatten each kinase matrix
ravel_kinase = list(range(len(IC_array)))
for w in range(len(IC_array)):
    ravel_kinase[w] = np.nan_to_num(np.ravel(IC_array[w]))
kinase_ravel_dict = {k:v for k,v in zip(kinases,ravel_kinase)}

#flatten each cluster matrix
ravel_cluster = list(range(len(IC_cluster_array)))
for q in range(len(IC_cluster_array)):
    ravel_cluster[q] = np.nan_to_num(np.ravel(IC_cluster_array[q]))
cluster_ravel_dict = {k:v for k,v in zip(clusters_name,ravel_cluster)}

#dot product calculations
dotscores = list(range(len(ravel_cluster)))
for u in range(len(ravel_cluster)):
    cluster_scores = list(range(len(ravel_kinase)))
    for p in range(len(ravel_kinase)):
        rank = np.dot(ravel_cluster[u],ravel_kinase[p])
        cluster_scores[p] = rank
    dotscores[u] = cluster_scores
dotscores_df = pd.DataFrame(np.transpose(dotscores), columns = clusters_name, index = kinases.tolist())       

fileloc4 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\kinase cluster ranking.xlsx"
from openpyxl import load_workbook
book = load_workbook(fileloc4)
writer = pd.ExcelWriter(fileloc4, engine = 'openpyxl')
writer.book = book
dotscores_df.to_excel(writer, sheet_name = 'dot product ranking')
writer.save()
writer.close()

#correlation coefficient calculations
corrscores = list(range(len(ravel_cluster)))
for u in range(len(ravel_cluster)):
    cluster_scores = list(range(len(ravel_kinase)))
    for p in range(len(ravel_kinase)):
        rank = np.corrcoef(ravel_cluster[u],ravel_kinase[p])[0][1]
        cluster_scores[p] = rank
    corrscores[u] = cluster_scores
corr_df = pd.DataFrame(np.transpose(corrscores), columns = clusters_name, index = kinases.tolist())       

fileloc4 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\kinase cluster ranking.xlsx"
from openpyxl import load_workbook
book = load_workbook(fileloc4)
writer = pd.ExcelWriter(fileloc4, engine = 'openpyxl')
writer.book = book
corr_df.to_excel(writer, sheet_name = 'pearson corr ranking')
writer.save()
writer.close()


#create a list of random peptides
import random
AA_noJ = ['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','Y']
randpep_list = list(range(12))
for r in range(len(randpep_list)):
    rand_AA = list(range(1000))
    for w in range(len(rand_AA)):
        base = np.random.choice(AA_noJ).tolist()
        rand_AA[w] = base
    randpep_list[r] = rand_AA
randpep = np.transpose(np.asarray(randpep_list))
randpep = pd.DataFrame(randpep)
ns = (randpep.groupby([0]).size()/randpep.groupby([0]).size().sum()).to_dict()
nf = (randpep.groupby([1]).size()/randpep.groupby([1]).size().sum()).to_dict()
nfr = (randpep.groupby([2]).size()/randpep.groupby([2]).size().sum()).to_dict()
nt = (randpep.groupby([3]).size()/randpep.groupby([3]).size().sum()).to_dict()
ntw = (randpep.groupby([4]).size()/randpep.groupby([4]).size().sum()).to_dict()
no = (randpep.groupby([5]).size()/randpep.groupby([5]).size().sum()).to_dict()
o = (randpep.groupby([6]).size()/randpep.groupby([6]).size().sum()).to_dict()
tw = (randpep.groupby([7]).size()/randpep.groupby([7]).size().sum()).to_dict()
t = (randpep.groupby([8]).size()/randpep.groupby([8]).size().sum()).to_dict()
fr = (randpep.groupby([9]).size()/randpep.groupby([9]).size().sum()).to_dict()
f = (randpep.groupby([10]).size()/randpep.groupby([10]).size().sum()).to_dict()
s = (randpep.groupby([11]).size()/randpep.groupby([11]).size().sum()).to_dict()
freqrand = [ns,nf,nfr,nt,ntw,no,o,tw,t,fr,f,s]
for j in freqrand: 
    for k in range(len(AA_noJ)):
        if AA_noJ[k] not in j:
            j[AA_noJ[k]] = 0  
M = 20
#columns
N = 12
init = [ [ 0 for i in range(M) ] for j in range(N) ]
for l in range(M): 
    for m in range(N):
        init[m][l] = freqrand[m][AA_noJ[l]]      
randpep_array =  np.transpose(np.asarray(init))
IC_randpep = randpep_array*np.log2(randpep_array/IMCDbkgd_array)
ravel_randpep = np.nan_to_num(np.ravel(IC_randpep))
corr_randpep = list(range(len(ravel_kinase)))
for y in range(len(ravel_kinase)):
    corr_randpep[y] = np.corrcoef(ravel_randpep,ravel_kinase[y])[0][1]
    
#sum of information content for positions
kinase_sums = []
for v in range(12):
    pos_sums = 0
    for b in range(len(IC_array)):
        pos_sums = pos_sums + sum(abs(np.nan_to_num(np.transpose(IC_array[b])[v])))
    kinase_sums.append(pos_sums)

import matplotlib.pyplot as plt
import seaborn as sns

kinase_sums.insert(6,0)
positions = ['-6','-5','-4','-3','-2','-1','0','+1','+2','+3','+4','+5','+6']
pos_weight = plt.bar(positions,kinase_sums, color = 'grey')
pos_weight = sns.set_style('white')
pos_weight = sns.set_style('ticks')
pos_weight = sns.despine()
plt.ylabel('Bits', fontsize = 18)
plt.xlabel('Position relative to phosphosite', fontsize = 18)
plt.tight_layout()
plt.savefig("J:\Depot - dDAVP-time course - Kirby\Figures\IC position weight",
            dpi=600)
    



#STY screening
STY_freq = list(range(len(freq_ind)))
for h in range(len(freq_ind)):
    STY_freq[h] = np.transpose(freq_ind[h])[6]

STY_freq_1 = list(range(len(clusters)))
for g in range(len(clusters)):
    STY_freq_1[g] = np.transpose(clusters[g])[6]
    
STY_dotscores = list(range(len(STY_freq_1)))
for u in range(len(STY_freq_1)):
    kinSTYscores = list(range(len(STY_freq)))
    for p in range(len(STY_freq)):
        rank = np.dot(STY_freq[p],STY_freq_1[u])
        kinSTYscores[p] = rank
    STY_dotscores[u] = kinSTYscores
STYdotscores_df = pd.DataFrame(np.transpose(STY_dotscores), columns = clusters_name, index = kinases.tolist())       


fileloc4 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\STY dot product ranking.xlsx"
from openpyxl import load_workbook
book = load_workbook(fileloc4)
writer = pd.ExcelWriter(fileloc4, engine = 'openpyxl')
writer.book = book
STYdotscores_df.to_excel(writer, sheet_name = 'imported ranking')
writer.save()
writer.close()













