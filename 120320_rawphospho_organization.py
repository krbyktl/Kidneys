# -*- coding: utf-8 -*-
"""
Created on Thu Dec  3 10:24:01 2020

@author: leokt

more organization of raw MaxQuant data
"""
#Aqp2 redo for Andromeda
import pandas as pd
fileloc1 = "J:\Depot - dDAVP-time course - Kirby\Analysis\\Aqp2_Msms.xlsx"
df_aqp2 = pd.read_excel(fileloc1, sheet_name="Aqp2_msms")

df_aqp2 = df_aqp2[~df_aqp2['Raw file'].str.contains('Total')]

df_TMT1 = df_aqp2[df_aqp2['Raw file'].str.contains('TMT1')]
df_TMT1 = df_TMT1.rename(columns = {'Reporter intensity corrected 1':'C1',
                         'Reporter intensity corrected 2':'C2',
                         'Reporter intensity corrected 3':'C3',
                         'Reporter intensity corrected 4':'C4',
                         'Reporter intensity corrected 8':'V1',
                         'Reporter intensity corrected 9':'V2',
                         'Reporter intensity corrected 10':'V3',
                         'Reporter intensity corrected 11':'V4'})
TMT1 = ['C1','C2','C3','C4','V1','V2','V3','V4']
norm_factor1 = [1.01, 0.91, 0.96, 1.03, 1, 0.95, 1, 1]
for i in range(len(TMT1)):
    df_TMT1[TMT1[i]] = norm_factor1[i]*df_TMT1[TMT1[i]]

df_TMT2 = df_aqp2[df_aqp2['Raw file'].str.contains('TMT2')]
df_TMT2 = df_TMT2.rename(columns = {'Reporter intensity corrected 1':'V1',
                         'Reporter intensity corrected 2':'V2',
                         'Reporter intensity corrected 3':'V3',
                         'Reporter intensity corrected 4':'V4',
                         'Reporter intensity corrected 8':'C1',
                         'Reporter intensity corrected 9':'C2',
                         'Reporter intensity corrected 10':'C3',
                         'Reporter intensity corrected 11':'C4'})
TMT2 = ['V1','V2','V3','V4','C1','C2','C3','C4']
norm_factor2 = [0.96, 1.05, 0.96, 1.13, 1, 1, 0.92, 1.07]
for i in range(len(TMT2)):
    df_TMT2[TMT2[i]] = norm_factor2[i]*df_TMT2[TMT2[i]]

df_TMT3 = df_aqp2[df_aqp2['Raw file'].str.contains('TMT3')]
df_TMT3 = df_TMT3.rename(columns = {'Reporter intensity corrected 1':'C4',
                         'Reporter intensity corrected 2':'C3',
                         'Reporter intensity corrected 3':'C2',
                         'Reporter intensity corrected 4':'C1',
                         'Reporter intensity corrected 8':'V4',
                         'Reporter intensity corrected 9':'V3',
                         'Reporter intensity corrected 10':'V2',
                         'Reporter intensity corrected 11':'V1'})
TMT3 = ['C4','C3','C2','C1','V4','V3','V2','V1']
norm_factor3 = [0.99, 1.20, 0.96, 0.95, 1.21, 1.01, 1, 1]
for i in range(len(TMT3)):
    df_TMT3[TMT3[i]] = norm_factor3[i]*df_TMT3[TMT3[i]]

df_aqp2_scale = pd.concat([df_TMT1, df_TMT2, df_TMT3])
monophos = df_aqp2_scale[df_aqp2_scale['Phospho (STY)'] == 1]
diphos = df_aqp2_scale[df_aqp2_scale['Phospho (STY)'] == 2]
triphos = df_aqp2_scale[df_aqp2_scale['Phospho (STY)'] == 3]

mono_uniq_seq = pd.unique(monophos['Modified sequence'])
di_uniq_seq = pd.unique(diphos['Modified sequence'])
tri_uniq_seq = pd.unique(triphos['Modified sequence'])


from openpyxl import load_workbook
fileloc1 = "J:\Depot - dDAVP-time course - Kirby\Analysis\\Aqp2_Msms_1.xlsx"
book = load_workbook(fileloc1)
writer = pd.ExcelWriter(fileloc1, engine = 'openpyxl')
writer.book = book
for i in mono_uniq_seq:
    monophos[monophos['Modified sequence'] == i].to_excel(writer,
                                                          sheet_name = str(list(mono_uniq_seq).index(i)) + '_1')
for j in di_uniq_seq:
    diphos[diphos['Modified sequence'] == j].to_excel(writer,
                                                      sheet_name = str(list(di_uniq_seq).index(j)) + '_2')
for k in tri_uniq_seq:
    triphos[triphos['Modified sequence'] == k].to_excel(writer,
                                                      sheet_name = str(list(tri_uniq_seq).index(k)) + '_3')

writer.save()
writer.close()











#Clip1 analysis
fileloc1 = "J:\Depot - dDAVP-time course - Kirby\Analysis\\Clip1_Msms.xlsx"
df_clip1 = pd.read_excel(fileloc1, sheet_name="Clip1_msms")

df_clip1 = df_clip1[~df_clip1['Raw file'].str.contains('Total')]

df_TMT1 = df_clip1[df_clip1['Raw file'].str.contains('TMT1')]
df_TMT1 = df_TMT1.rename(columns = {'Reporter intensity corrected 1':'C1',
                         'Reporter intensity corrected 2':'C2',
                         'Reporter intensity corrected 3':'C3',
                         'Reporter intensity corrected 4':'C4',
                         'Reporter intensity corrected 8':'V1',
                         'Reporter intensity corrected 9':'V2',
                         'Reporter intensity corrected 10':'V3',
                         'Reporter intensity corrected 11':'V4'})
TMT1 = ['C1','C2','C3','C4','V1','V2','V3','V4']
norm_factor1 = [1.01, 0.91, 0.96, 1.03, 1, 0.95, 1, 1]
for i in range(len(TMT1)):
    df_TMT1[TMT1[i]] = norm_factor1[i]*df_TMT1[TMT1[i]]

df_TMT2 = df_clip1[df_clip1['Raw file'].str.contains('TMT2')]
df_TMT2 = df_TMT2.rename(columns = {'Reporter intensity corrected 1':'V1',
                         'Reporter intensity corrected 2':'V2',
                         'Reporter intensity corrected 3':'V3',
                         'Reporter intensity corrected 4':'V4',
                         'Reporter intensity corrected 8':'C1',
                         'Reporter intensity corrected 9':'C2',
                         'Reporter intensity corrected 10':'C3',
                         'Reporter intensity corrected 11':'C4'})
TMT2 = ['V1','V2','V3','V4','C1','C2','C3','C4']
norm_factor2 = [0.96, 1.05, 0.96, 1.13, 1, 1, 0.92, 1.07]
for i in range(len(TMT2)):
    df_TMT2[TMT2[i]] = norm_factor2[i]*df_TMT2[TMT2[i]]

df_TMT3 = df_clip1[df_clip1['Raw file'].str.contains('TMT3')]
df_TMT3 = df_TMT3.rename(columns = {'Reporter intensity corrected 1':'C4',
                         'Reporter intensity corrected 2':'C3',
                         'Reporter intensity corrected 3':'C2',
                         'Reporter intensity corrected 4':'C1',
                         'Reporter intensity corrected 8':'V4',
                         'Reporter intensity corrected 9':'V3',
                         'Reporter intensity corrected 10':'V2',
                         'Reporter intensity corrected 11':'V1'})
TMT3 = ['C4','C3','C2','C1','V4','V3','V2','V1']
norm_factor3 = [0.99, 1.20, 0.96, 0.95, 1.21, 1.01, 1, 1]
for i in range(len(TMT3)):
    df_TMT3[TMT3[i]] = norm_factor3[i]*df_TMT3[TMT3[i]]

df_clip1_scale = pd.concat([df_TMT1, df_TMT2, df_TMT3])
monophos = df_clip1_scale[df_clip1_scale['Phospho (STY)'] == 1]
diphos = df_clip1_scale[df_clip1_scale['Phospho (STY)'] == 2]
triphos = df_clip1_scale[df_clip1_scale['Phospho (STY)'] == 3]

mono_uniq_seq = pd.unique(monophos['Modified sequence'])
di_uniq_seq = pd.unique(diphos['Modified sequence'])

from openpyxl import load_workbook
book = load_workbook(fileloc1)
writer = pd.ExcelWriter(fileloc1, engine = 'openpyxl')
writer.book = book
for i in mono_uniq_seq:
    monophos[monophos['Modified sequence'] == i].to_excel(writer,
                                                          sheet_name = str(list(mono_uniq_seq).index(i)) + '_1')
for j in di_uniq_seq:
    diphos[diphos['Modified sequence'] == j].to_excel(writer,
                                                      sheet_name = str(list(di_uniq_seq).index(j)) + '_2')
writer.save()
writer.close()
               

    
#Slc14a2
import pandas as pd
fileloc2 = "J:\Depot - dDAVP-time course - Kirby\Analysis\\Slc14a2_Msms.xlsx"
df_slc14a2 = pd.read_excel(fileloc2, sheet_name="Slc14a2_msms")

df_slc14a2 = df_slc14a2[~df_slc14a2['Raw file'].str.contains('Total')]

df_TMT1 = df_slc14a2[df_slc14a2['Raw file'].str.contains('TMT1')]
df_TMT1 = df_TMT1.rename(columns = {'Reporter intensity corrected 1':'C1',
                         'Reporter intensity corrected 2':'C2',
                         'Reporter intensity corrected 3':'C3',
                         'Reporter intensity corrected 4':'C4',
                         'Reporter intensity corrected 8':'V1',
                         'Reporter intensity corrected 9':'V2',
                         'Reporter intensity corrected 10':'V3',
                         'Reporter intensity corrected 11':'V4'})
TMT1 = ['C1','C2','C3','C4','V1','V2','V3','V4']
norm_factor1 = [1.01, 0.91, 0.96, 1.03, 1, 0.95, 1, 1]
for i in range(len(TMT1)):
    df_TMT1[TMT1[i]] = norm_factor1[i]*df_TMT1[TMT1[i]]

df_TMT2 = df_slc14a2[df_slc14a2['Raw file'].str.contains('TMT2')]
df_TMT2 = df_TMT2.rename(columns = {'Reporter intensity corrected 1':'V1',
                         'Reporter intensity corrected 2':'V2',
                         'Reporter intensity corrected 3':'V3',
                         'Reporter intensity corrected 4':'V4',
                         'Reporter intensity corrected 8':'C1',
                         'Reporter intensity corrected 9':'C2',
                         'Reporter intensity corrected 10':'C3',
                         'Reporter intensity corrected 11':'C4'})
TMT2 = ['V1','V2','V3','V4','C1','C2','C3','C4']
norm_factor2 = [0.96, 1.05, 0.96, 1.13, 1, 1, 0.92, 1.07]
for i in range(len(TMT2)):
    df_TMT2[TMT2[i]] = norm_factor2[i]*df_TMT2[TMT2[i]]

df_TMT3 = df_slc14a2[df_slc14a2['Raw file'].str.contains('TMT3')]
df_TMT3 = df_TMT3.rename(columns = {'Reporter intensity corrected 1':'C4',
                         'Reporter intensity corrected 2':'C3',
                         'Reporter intensity corrected 3':'C2',
                         'Reporter intensity corrected 4':'C1',
                         'Reporter intensity corrected 8':'V4',
                         'Reporter intensity corrected 9':'V3',
                         'Reporter intensity corrected 10':'V2',
                         'Reporter intensity corrected 11':'V1'})
TMT3 = ['C4','C3','C2','C1','V4','V3','V2','V1']
norm_factor3 = [0.99, 1.20, 0.96, 0.95, 1.21, 1.01, 1, 1]
for i in range(len(TMT3)):
    df_TMT3[TMT3[i]] = norm_factor3[i]*df_TMT3[TMT3[i]]

df_slc14a2_scale = pd.concat([df_TMT1, df_TMT2, df_TMT3])
monophos = df_slc14a2_scale[df_slc14a2_scale['Phospho (STY)'] == 1]
diphos = df_slc14a2_scale[df_slc14a2_scale['Phospho (STY)'] == 2]
triphos = df_slc14a2_scale[df_slc14a2_scale['Phospho (STY)'] == 3]

mono_uniq_seq = pd.unique(monophos['Modified sequence'])
di_uniq_seq = pd.unique(diphos['Modified sequence'])
tri_uniq_seq = pd.unique(triphos['Modified sequence'])


from openpyxl import load_workbook
book = load_workbook(fileloc2)
writer = pd.ExcelWriter(fileloc2, engine = 'openpyxl')
writer.book = book
for i in mono_uniq_seq:
    monophos[monophos['Modified sequence'] == i].to_excel(writer,
                                                          sheet_name = str(list(mono_uniq_seq).index(i)) + '_1')
for j in di_uniq_seq:
    diphos[diphos['Modified sequence'] == j].to_excel(writer,
                                                      sheet_name = str(list(di_uniq_seq).index(j)) + '_2')
for k in tri_uniq_seq:
    triphos[triphos['Modified sequence'] == k].to_excel(writer,
                                                      sheet_name = str(list(tri_uniq_seq).index(k)) + '_3')

writer.save()
writer.close()
