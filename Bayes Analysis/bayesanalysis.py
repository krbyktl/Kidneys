# -*- coding: utf-8 -*-
"""
Created on Tue May 11 18:24:37 2021

@author: leokt
"""

# =============================================================================
# Steps:

# pre-analyzed:
# IMCD microarray transcriptome (Uawithya et al 2008)
# IMCD RNA-seq transcriptome (Lee et al 2015)

# mouse IMCD RNA seq (Chen et al 2021)
# IMCD suspension proteome (data collected in this paper)
# IMCD microdissected proteome (Limbutara et al 2020)
#
# Screening for S/T kinases (Sugiyama et al 2019)
# Colocalization from fractions (Yang et al 2015)
# Position ranking (Sugiyama et al 2019) (x12 for each position)
# Known kinases
# =============================================================================

import pandas as pd
import numpy as np
from bayesmodules import interpolate
from bayesmodules import cMBF_calc


fileloc_1 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\expression data.xlsx"
express_df = pd.read_excel(fileloc_1, None)

#from sugiyama_data import STYdotscores_df
main_vec = express_df['exp bayes calc V2']
STYdots = interpolate(STYdotscores_df)
main_vec = main_vec.rename(columns = {'Gene Symbol':'Kinases'})

# %%
#STY ranking
main_vec = main_vec.merge(STYdots,on='Kinases',how='left')
clus_names = STYdots.columns.tolist()
clus_names.remove('Kinases')
for i in clus_names: 
    STY_cMBF = cMBF_calc(main_vec[i],1/3)
    prod = STY_cMBF*main_vec['Posterior_Exp']
    main_vec[i + '_STY'] = prod/sum(prod)
main_vec = main_vec.drop(clus_names,axis=1)

fileloc_3 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\BayesTC complete calc_python V5.xlsx"
writer=pd.ExcelWriter(fileloc_3)
main_vec.to_excel(writer, sheet_name='STY ranking')

writer.save()

# %%
#colocalization ranking
fileloc_2 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\colocalization mapping V2.xlsx"
coloc_df = pd.read_excel(fileloc_2, "coloc dot ranking")

coloc_vec = main_vec
coloc_vec = coloc_vec.merge(coloc_df,on='Kinases',how='left')
for i in clus_names:
    coloc_cMBF = cMBF_calc(coloc_vec[i],1/3)
    prod1 = coloc_cMBF*main_vec[i + '_STY']
    coloc_vec[i + '_coloc'] = prod1/sum(prod1)
coloc_vec = coloc_vec.drop(clus_names,axis=1)

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
coloc_vec.to_excel(writer, sheet_name = 'coloc ranking')
writer.save()
writer.close()

# %%
#position ranking
#from sugiyama_data import pos_dotscore_list
posit = ['-6','-5','-4','-3','-2','-1','+1','+2','+3','+4','+5','+6']

position_based = main_vec[['Kinases']].copy()
int_pos = list(range(len(pos_dotscore_list)))
for t in range(len(pos_dotscore_list)):
    int_pos[t] = interpolate(pos_dotscore_list[t])


# %%
# only position +1
int_pos6 = int_pos[6]  
variable = coloc_vec.merge(int_pos6,on='Kinases',how='left')
for i in clus_names: 
    pos_cMBF = cMBF_calc(variable[i],max(variable[i])/12)
    prod = pos_cMBF*variable[i + '_coloc']
    position_based[i + '_+1'] = prod/sum(prod)
    
from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
position_based.to_excel(writer, sheet_name = 'IC position ranking +1')
writer.save()
writer.close()

# %%
# position -2 following

position_based_2 = main_vec[['Kinases']].copy()
int_pos4 = int_pos[4]  
variable = position_based.merge(int_pos4,on='Kinases',how='left')
for i in clus_names: 
    pos_cMBF = cMBF_calc(variable[i],max(variable[i])/12)
    prod = pos_cMBF*variable[i + '_+1']
    position_based_2[i + '_-2'] = prod/sum(prod)

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
position_based_2.to_excel(writer, sheet_name = 'IC position ranking -2')
writer.save()
writer.close()

# %%
# position -1 following

position_based_3 = main_vec[['Kinases']].copy()
int_pos5 = int_pos[5]  
variable = position_based_2.merge(int_pos5,on='Kinases',how='left')
for i in clus_names: 
    pos_cMBF = cMBF_calc(variable[i],max(variable[i])/12)
    prod = pos_cMBF*variable[i + '_-2']
    position_based_3[i + '_-1'] = prod/sum(prod)

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
position_based_3.to_excel(writer, sheet_name = 'IC position ranking -1')
writer.save()
writer.close()

# %%
# position -3 following

position_based_4 = main_vec[['Kinases']].copy()
int_pos3 = int_pos[3]  
variable = position_based_3.merge(int_pos3,on='Kinases',how='left')
for i in clus_names: 
    pos_cMBF = cMBF_calc(variable[i],max(variable[i])/12)
    prod = pos_cMBF*variable[i + '_-1']
    position_based_4[i + '_-3'] = prod/sum(prod)

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
position_based_4.to_excel(writer, sheet_name = 'IC position ranking -3')
writer.save()
writer.close()

# %%
# position +2 following

position_based_5 = main_vec[['Kinases']].copy()
int_pos7 = int_pos[7]  
variable = position_based_4.merge(int_pos7,on='Kinases',how='left')
for i in clus_names: 
    pos_cMBF = cMBF_calc(variable[i],max(variable[i])/12)
    prod = pos_cMBF*variable[i + '_-3']
    position_based_5[i + '_+2'] = prod/sum(prod)

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
position_based_5.to_excel(writer, sheet_name = 'IC position ranking +2 12th')
writer.save()
writer.close()

# %%    
# entire position ranking
for c in range(len(clus_names)):
    new_prior = coloc_vec[clus_names[c] + '_coloc']
    for p in range(len(int_pos)):
        variable = coloc_vec.merge(int_pos[p],on='Kinases',how='left')
        pos_cMBF = cMBF_calc(variable[clus_names[c]],max(variable[clus_names[c]])/10)
        prod2 = pos_cMBF*new_prior
        new_prior = prod2/sum(prod2)
    position_based[clus_names[c]+'_pos'] = new_prior.tolist()

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
position_based.to_excel(writer, sheet_name = 'IC position ranking')
writer.save()
writer.close()

# %%
# known kinases
fileloc_4 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\Bayes sheets\\BayesTC_known kinases.xlsx"
known_df = pd.read_excel(fileloc_4, "Simplified combined")

cluster_direct = ['Increase','Decrease','Decrease','Decrease','Decrease','Increase','Increase','Decrease',
                  'Decrease','Increase','Decrease','Decrease','Increase','Decrease']

known_based = main_vec[['Kinases']].copy()
for a in range(len(clus_names)):
    new_likelihood = list(range(len(position_based['Kinases'])))
    for b in range(len(position_based['Kinases'])):
        if position_based['Kinases'][b] in known_df['Kinases'].values:
            if known_df[known_df['Kinases']==position_based['Kinases'][b]]['Net Effect on Activity'].tolist()[0] == cluster_direct[a]:
                new_likelihood[b] = 0.9
            else:
                new_likelihood[b] = 0.1
        else:
            new_likelihood[b] = 0.5
    known_based[clus_names[a]] = new_likelihood

known_vec = main_vec[['Kinases']].copy()
for c in clus_names:
    prod2 = position_based[c+'_pos']*known_based[c]
    new_prior = prod2/sum(prod2)
    known_vec[c+'_known'] = new_prior.tolist()
    
from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
known_vec.to_excel(writer, sheet_name = 'known kinase ranking')
writer.save()
writer.close()






