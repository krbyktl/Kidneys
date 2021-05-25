# -*- coding: utf-8 -*-
"""
Created on Fri May 21 20:40:47 2021

@author: leokt
"""

# =============================================================================
# Steps:

# pre-analyzed:
# mouse IMCD RNA seq (Chen et al 2021)
# IMCD suspension proteome (data collected in this paper)
# IMCD microdissected proteome (Limbutara et al 2020)
#
# Screening for S/T kinases (Sugiyama et al 2019)
# Colocalization from fractions (Yang et al 2015)
# Known kinases
# ranking based on kinase family group
# =============================================================================


import pandas as pd
import numpy as np
from bayesmodules import interpolate
from bayesmodules import cMBF_calc


fileloc_1 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\expression data.xlsx"
express_df = pd.read_excel(fileloc_1, None)

#from sugiyama_data import STYdotscores_df
main_vec = express_df['exp bayes calc V2']
STYdots = interpolate(STYdotscores_df)
main_vec = main_vec.rename(columns = {'Gene Symbol':'Kinases'})

# %%
#STY ranking
main_vec = main_vec.merge(STYdots,on='Kinases',how='left')
clus_names = STYdots.columns.tolist()
clus_names.remove('Kinases')
for i in clus_names: 
    STY_cMBF = cMBF_calc(main_vec[i],1/9)
    prod = STY_cMBF*main_vec['Posterior_Exp']
    main_vec[i + '_STY'] = prod/sum(prod)
main_vec = main_vec.drop(clus_names,axis=1)

fileloc_3 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\BayesTC complete calc_python V13.xlsx"
writer=pd.ExcelWriter(fileloc_3)
main_vec.to_excel(writer, sheet_name='STY ranking')

writer.save()

# %%
#colocalization ranking
fileloc_2 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\colocalization mapping V2.xlsx"
coloc_df = pd.read_excel(fileloc_2, "coloc dot ranking V2")

coloc_vec = main_vec
coloc_vec = coloc_vec.merge(coloc_df,on='Kinases',how='left')
for i in clus_names:
    coloc_cMBF = cMBF_calc(coloc_vec[i],np.mean(coloc_vec[i]))
    prod1 = coloc_cMBF*main_vec[i + '_STY']
    coloc_vec[i + '_coloc'] = prod1/sum(prod1)
coloc_vec = coloc_vec.drop(clus_names,axis=1)

from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
coloc_vec.to_excel(writer, sheet_name = 'coloc ranking')
writer.save()
writer.close()

# %%
# known kinases
fileloc_4 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\Bayes sheets\\BayesTC_known kinases.xlsx"
known_df = pd.read_excel(fileloc_4, "Simplified combined")

cluster_direct = ['Increase','Decrease','Decrease','Decrease','Decrease','Increase','Increase','Decrease',
                  'Decrease','Increase','Decrease','Decrease','Increase','Decrease']

clus_names = STYdots.columns.tolist()
clus_names.remove('Kinases')
position_based = coloc_vec
known_based = coloc_vec[['Kinases']].copy()
for a in range(len(clus_names)):
    new_likelihood = list(range(len(position_based['Kinases'])))
    for b in range(len(position_based['Kinases'])):
        if position_based['Kinases'][b] in known_df['Kinases'].values:
            if known_df[known_df['Kinases']==position_based['Kinases'][b]]['Net Effect on Activity'].tolist()[0] == cluster_direct[a]:
                new_likelihood[b] = 0.9
            else:
                new_likelihood[b] = 0.1
        else:
            new_likelihood[b] = 0.5
    known_based[clus_names[a]] = new_likelihood

known_vec = main_vec[['Kinases']].copy()
for c in clus_names:
    prod2 = position_based[c + "_coloc"]*known_based[c]
    new_prior = prod2/sum(prod2)
    known_vec[c+'_known'] = new_prior.tolist()
    
from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
known_vec.to_excel(writer, sheet_name = 'known kinase ranking')
writer.save()
writer.close()

# %%
# incorporating family

fileloc_5 = "J:\Depot - dDAVP-time course - Kirby\BayesAnalysis\\kinase family.xlsx"
fam_df = pd.read_excel(fileloc_5, "family assignment")
baso_df = pd.read_excel(fileloc_5, "Basophilic assignment")
pro_df = pd.read_excel(fileloc_5, "Proline assignment")

fam_vec = known_vec.merge(fam_df,on='Kinases',how='left')

pred_baso = ['IIA1a','IIA1b','IIB1a','IIB1b']
baso_assign = fam_vec.merge(baso_df,on='Family',how='left')
clus_names = STYdots.columns.tolist()
clus_names.remove('Kinases')
baso_vec = known_vec[['Kinases']].copy()
for d in range(len(clus_names)):
    if clus_names[d] in pred_baso:
        prod = baso_assign[clus_names[d] + '_known']*baso_assign['Probability']
        prod.fillna(value = 0, 
          inplace = True)
        new_prior = prod/sum(prod)
        baso_vec[clus_names[d] + '_rank_F'] = new_prior.tolist()
    else:
        baso_vec[clus_names[d] + '_norank'] = baso_assign[clus_names[d] + '_known']
   
from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
baso_vec.to_excel(writer, sheet_name = 'baso family')
writer.save()
writer.close()

pred_pro = ['IB','IA1','IA2a','IA2b','IA2c']
pro_assign = fam_vec.merge(pro_df,on='Family',how='left')
pro_vec = known_vec[['Kinases']].copy()
for e in range(len(clus_names)):
    if clus_names[e] in pred_pro:
        prod = pro_assign[clus_names[e] + '_known']*pro_assign['Probability']
        prod.fillna(value = 0, 
          inplace = True)
        new_prior = prod/sum(prod)
        pro_vec[clus_names[e] + '_rank_F'] = new_prior.tolist()
    else:
        pro_vec[clus_names[e] + '_norank'] = pro_assign[clus_names[e] + '_known']
    
from openpyxl import load_workbook
book = load_workbook(fileloc_3)
writer = pd.ExcelWriter(fileloc_3, engine = 'openpyxl')
writer.book = book
pro_vec.to_excel(writer, sheet_name = 'proline family')
writer.save()
writer.close()


